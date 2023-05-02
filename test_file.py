import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
import xlrd
import csv
import time
import os.path
import requests
from selenium import webdriver
from selene import browser
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager



# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_download_xlsx():
    file_xlsx = os.path.dirname(os.path.abspath('__file__'))
    workbook_path= os.path.join(file_xlsx, 'resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    value = sheet.cell(row=3, column=2).value
    assert value == 'Mara'




# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_book():

    book = xlrd.open_workbook('resources/file_example_XLS_10.xls')
print(f'Количество листов {book.nsheets}')
print(f'Имена листов {book.sheet_names()}')
sheet = book.sheet_by_index(0)
print(f'Количество столбцов {sheet.ncols}')
print(f'Количество строк {sheet.nrows}')
print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
# печать всех строк по очереди
for rx in range(sheet.nrows):
    print(sheet.row(rx))




# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_reader():


    reader = PdfReader("resources/docs-pytest-org-en-latest.pdf")
number_of_pages = len(reader.pages)
page = reader.pages[0]
text = page.extract_text()
print(page)
print(number_of_pages)
print(text)


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'

    r = requests.get(url)
    with open('selenium_logo.png', 'wb') as file:
        file.write(r.content)

    size = os.path.getsize('selenium_logo.png')

    assert size == 30803



# TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp

def test_download():
options = webdriver.ChromeOptions()
prefs = {
    "download.default_directory": '/Users/kot/GitHubProjects/qa-guru/qa_guru_python_5_7_files',
    "download.prompt_for_download": False
}
options.add_experimental_option("prefs", prefs)

browser.config.driver_options = options

browser.open("https://github.com/pytest-dev/pytest")
browser.element(".d-none .Button-label").click()
browser.element('[data-open-app="link"]').click()



# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_csv():
with open('resources/eggs.csv', 'w') as csvfile:
    csvwriter = csv.writer(csvfile, delimiter=',')
    csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
    csvwriter.writerow(['Alex', 'Serj', 'Yana'])

with open('resources/eggs.csv') as csvfile:
    csvreader = csv.reader(csvfile)
    for row in csvreader:
        print(row)